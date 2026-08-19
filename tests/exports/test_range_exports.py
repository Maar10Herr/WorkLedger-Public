import sqlite3
import zipfile
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from openpyxl import load_workbook

from apps.evidence.models import AttachmentLink
from apps.evidence.services import store_attachment
from apps.expenses.services import create_expense
from apps.exports.services import (
    build_full_zip,
    build_range_csv,
    build_range_json,
    build_range_sqlite,
    build_range_xlsx,
)
from apps.ledger.services import revise_event
from apps.taxes.journey import record_route_distance
from apps.travel.models import Location, LocationType, TransportMode
from apps.travel.services import create_journey

pytestmark = pytest.mark.django_db

RECORDED_AT = datetime(2026, 8, 5, 12, 0, tzinfo=UTC)
RECORDED_AFTER_AS_OF = datetime(2026, 8, 5, 12, 1, tzinfo=UTC)


@pytest.fixture(autouse=True)
def freeze_event_recording_time(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(timezone, "now", lambda: RECORDED_AT)


def make_expense_with_receipt(settings: Any, tmp_path: Path) -> None:
    settings.DATA_DIR = tmp_path / "data"
    expense = create_expense(
        effective_at=datetime(2026, 8, 4, 12, 0, tzinfo=UTC),
        category="train_ticket",
        amount=Decimal("42.50"),
        currency="EUR",
        tax_relevant=True,
        employer_reimbursable=True,
        employer_paid=False,
    )
    attachment = store_attachment(
        SimpleUploadedFile(
            "ticket.png", b"\x89PNG\r\n\x1a\n" + b"evidence", content_type="image/png"
        )
    )
    AttachmentLink.objects.create(
        attachment=attachment, event=expense.event, link_type="expense_receipt"
    )


def test_xlsx_export_is_deterministic_and_has_separate_tracks(
    settings: Any, tmp_path: Path
) -> None:
    make_expense_with_receipt(settings, tmp_path)
    as_of = RECORDED_AT
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"

    build_range_xlsx(date(2026, 8, 1), date(2026, 8, 31), as_of, first)
    build_range_xlsx(date(2026, 8, 1), date(2026, 8, 31), as_of, second)

    assert first.read_bytes() == second.read_bytes()
    workbook = load_workbook(first, data_only=False)
    assert {
        "Events",
        "Journeys",
        "Expenses",
        "PerDiem",
        "TaxOutput",
        "EmployerClaims",
        "Attachments",
        "Revisions",
        "TaxRules",
        "README",
    }.issubset(workbook.sheetnames)
    assert workbook["TaxOutput"]["F2"].value == 0
    assert workbook["EmployerClaims"]["E2"].value == Decimal("42.50")


def test_portable_sqlite_and_full_zip_include_audit_and_originals(
    settings: Any, tmp_path: Path
) -> None:
    make_expense_with_receipt(settings, tmp_path)
    as_of = RECORDED_AT
    sqlite_path = tmp_path / "ledger.sqlite3"
    zip_path = tmp_path / "complete.zip"

    build_range_sqlite(date(2026, 8, 1), date(2026, 8, 31), as_of, sqlite_path)
    build_full_zip(date(2026, 8, 1), date(2026, 8, 31), as_of, zip_path)

    with sqlite3.connect(sqlite_path) as connection:
        assert connection.execute("SELECT COUNT(*) FROM events").fetchone()[0] == 2
        assert connection.execute("SELECT COUNT(*) FROM revisions").fetchone()[0] == 2
        assert connection.execute("SELECT COUNT(*) FROM attachments").fetchone()[0] == 1
    with zipfile.ZipFile(zip_path) as archive:
        names = archive.namelist()
        assert {
            "workledger.xlsx",
            "workledger.sqlite3",
            "manifest.json",
            "data.json",
            "readme.md",
            "data-dictionary.md",
            "unresolved-items.md",
            "events.csv",
            "journeys.csv",
            "external-activities.csv",
            "derived-days.csv",
            "expenses.csv",
            "reimbursements.csv",
            "attachments.csv",
            "revisions.csv",
        }.issubset(names)
        assert any(
            name.startswith("attachments/originals/") and name.endswith("ticket.png")
            for name in names
        )


def test_as_of_export_uses_historical_revision_and_track_flags(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    expense = create_expense(
        effective_at=datetime(2026, 8, 4, 12, 0, tzinfo=UTC),
        category="train_ticket",
        amount=Decimal("42.50"),
        currency="EUR",
        tax_relevant=True,
        employer_reimbursable=True,
        employer_paid=False,
    )
    event = expense.event
    first_revision = event.current_revision
    assert first_revision is not None
    as_of = first_revision.recorded_at
    monkeypatch.setattr(timezone, "now", lambda: RECORDED_AFTER_AS_OF)
    changed = dict(first_revision.snapshot)
    changed.update(
        {
            "amount": "999.00",
            "tax_relevant": False,
            "employer_reimbursable": False,
        }
    )
    revise_event(
        event=event,
        effective_at=first_revision.effective_at,
        snapshot=changed,
        complete=True,
        comment="later correction",
    )
    event.refresh_from_db()
    assert event.tax_relevant is False
    assert event.employer_reimbursable is False

    destination = tmp_path / "historical.xlsx"
    build_range_xlsx(date(2026, 8, 1), date(2026, 8, 31), as_of, destination)
    workbook = load_workbook(destination, data_only=False)
    expense_row = next(workbook["Expenses"].iter_rows(min_row=2, values_only=True))
    assert expense_row[4] == "42.50"
    assert expense_row[7] is True
    assert expense_row[8] is True


def test_csv_and_json_variants_are_deterministic(tmp_path: Path) -> None:
    create_expense(
        effective_at=datetime(2026, 8, 4, 10, tzinfo=UTC),
        category="taxi",
        amount=Decimal("12.00"),
        currency="EUR",
        tax_relevant=True,
        employer_reimbursable=False,
        employer_paid=False,
    )
    as_of = RECORDED_AT
    for builder, suffix in ((build_range_csv, "csv"), (build_range_json, "json")):
        first = builder(date(2026, 8, 1), date(2026, 8, 31), as_of, tmp_path / f"a.{suffix}")
        second = builder(date(2026, 8, 1), date(2026, 8, 31), as_of, tmp_path / f"b.{suffix}")
        assert first.read_bytes() == second.read_bytes()


def test_non_car_commuting_allowance_is_capped_at_4500_eur(tmp_path: Path) -> None:
    home = Location.objects.create(
        name="Home", location_type=LocationType.RESIDENCE, is_default_residence=True
    )
    office = Location.objects.create(name="Office", location_type=LocationType.FIRST_WORKPLACE)
    record_route_distance(
        origin=home,
        destination=office,
        mode="driving",
        distance_km=Decimal("20000"),
        source="manual",
    )
    create_journey(
        origin=home,
        destination=office,
        transport_mode=TransportMode.TRAIN,
        effective_at=datetime(2026, 8, 4, 8, tzinfo=UTC),
        tax_relevant=True,
    )
    output = build_range_xlsx(
        date(2026, 8, 1),
        date(2026, 8, 31),
        RECORDED_AT,
        tmp_path / "cap.xlsx",
    )
    workbook = load_workbook(output, data_only=False)
    assert workbook["TaxOutput"]["F2"].value == 4500


def test_events_recorded_after_as_of_are_excluded(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    create_expense(
        effective_at=datetime(2026, 8, 4, 10, tzinfo=UTC),
        category="taxi",
        amount=Decimal("12.00"),
        currency="EUR",
        tax_relevant=True,
        employer_reimbursable=False,
        employer_paid=False,
    )
    monkeypatch.setattr(timezone, "now", lambda: RECORDED_AFTER_AS_OF)
    create_expense(
        effective_at=datetime(2026, 8, 4, 11, tzinfo=UTC),
        category="train_ticket",
        amount=Decimal("42.50"),
        currency="EUR",
        tax_relevant=True,
        employer_reimbursable=False,
        employer_paid=False,
    )

    output = build_range_csv(
        date(2026, 8, 1), date(2026, 8, 31), RECORDED_AT, tmp_path / "as-of.csv"
    )

    assert len(output.read_text(encoding="utf-8").splitlines()) == 2
