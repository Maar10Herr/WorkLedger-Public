from __future__ import annotations

import csv
import hashlib
import io
import json
import sqlite3
import tempfile
import uuid
import zipfile
from collections.abc import Callable
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING, Any

from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from apps.evidence.models import Attachment, AttachmentLink
from apps.expenses.services import expense_track_amounts
from apps.ledger.models import Event, EventRevision
from apps.taxes.journey import derive_journey_tax
from apps.taxes.models import PerDiemCalculation, TaxRule

if TYPE_CHECKING:
    from .models import ExportArtifact

FIXED_ZIP_TIME = (1980, 1, 1, 0, 0, 0)
SHEET_NAMES = (
    "Events",
    "Journeys",
    "ExternalActivities",
    "Expenses",
    "PerDiem",
    "TaxOutput",
    "EmployerClaims",
    "Reconciliation",
    "Attachments",
    "Revisions",
    "TaxRules",
    "README",
)


@dataclass(frozen=True)
class ExportEvent:
    event: Event
    revision: EventRevision


def _canonical(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _events_as_of(start: date, end: date, as_of: datetime) -> list[ExportEvent]:
    rows: list[ExportEvent] = []
    for event in Event.objects.all().order_by("id"):
        revision = event.revisions.filter(recorded_at__lte=as_of).order_by(
            "-recorded_at", "-revision_number"
        ).first()
        if revision is None or revision.deleted:
            continue
        if start <= revision.effective_at.date() <= end:
            rows.append(ExportEvent(event, revision))
    rows.sort(key=lambda row: (row.revision.effective_at, str(row.event.pk)))
    return rows


def _write_sheet(workbook: Workbook, name: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="16624F")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        values = [str(cell.value or "") for cell in column]
        width = min(60, max(10, max(map(len, values)) + 2))
        sheet.column_dimensions[column[0].column_letter].width = width


def _normalized_zip(source: Path, destination: Path) -> None:
    with zipfile.ZipFile(source, "r") as incoming, zipfile.ZipFile(
        destination, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as outgoing:
        for name in sorted(incoming.namelist()):
            info = zipfile.ZipInfo(name, FIXED_ZIP_TIME)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o100644 << 16
            outgoing.writestr(info, incoming.read(name), compress_type=zipfile.ZIP_DEFLATED)


def _tax_rule_rows(start: date, end: date) -> list[TaxRule]:
    return list(
        TaxRule.objects.filter(effective_from__lte=end)
        .order_by("effective_from", "code")
    )


def build_range_xlsx(
    start: date,
    end: date,
    as_of: datetime,
    destination: Path,
    events: list[ExportEvent] | None = None,
) -> Path:
    selected = events if events is not None else _events_as_of(start, end, as_of)
    workbook = Workbook()
    active_sheet = workbook.active
    assert active_sheet is not None
    workbook.remove(active_sheet)
    workbook.properties.creator = "WorkLedger"
    workbook.properties.lastModifiedBy = "WorkLedger"
    workbook.properties.created = datetime(1980, 1, 1)
    workbook.properties.modified = datetime(1980, 1, 1)

    event_rows = [
        [
            str(row.event.pk),
            row.event.event_type,
            row.revision.effective_at.isoformat(),
            row.revision.recorded_at.isoformat(),
            row.revision.revision_number,
            row.revision.complete,
            bool(row.revision.snapshot.get("tax_relevant", row.event.tax_relevant)),
            bool(
                row.revision.snapshot.get(
                    "employer_reimbursable", row.event.employer_reimbursable
                )
            ),
            _canonical(row.revision.snapshot),
            row.revision.audit_hash,
        ]
        for row in selected
    ]
    _write_sheet(
        workbook,
        "Events",
        [
            "event_id",
            "event_type",
            "effective_at",
            "recorded_at",
            "revision",
            "complete",
            "tax_relevant",
            "employer_reimbursable",
            "facts_json",
            "audit_hash",
        ],
        event_rows,
    )

    journey_rows: list[list[Any]] = []
    external_activity_rows: list[list[Any]] = []
    expense_rows: list[list[Any]] = []
    tax_rows: list[list[Any]] = []
    employer_rows: list[list[Any]] = []
    commute_days: set[tuple[date, str, str]] = set()
    capped_commuting_totals: dict[int, Decimal] = {}
    for row in selected:
        snapshot = row.revision.snapshot
        tax_relevant = bool(snapshot.get("tax_relevant", row.event.tax_relevant))
        employer_reimbursable = bool(
            snapshot.get("employer_reimbursable", row.event.employer_reimbursable)
        )
        if row.event.event_type == "journey":
            derivation = derive_journey_tax(row.event, row.revision, as_of)
            journey_rows.append(
                [
                    str(row.event.pk),
                    row.revision.effective_at.isoformat(),
                    snapshot.get("origin_name"),
                    snapshot.get("destination_name"),
                    snapshot.get("transport_mode"),
                    snapshot.get("actual_kilometres"),
                    derivation.classification,
                    derivation.rule_code,
                    str(derivation.amount),
                    derivation.complete,
                    snapshot.get("covered_by_pass"),
                    snapshot.get("rail_pass_name"),
                    snapshot.get("total_fare"),
                    snapshot.get("personally_paid"),
                    snapshot.get("employer_reimbursement"),
                    snapshot.get("employer_paid"),
                    snapshot.get("payer_description"),
                ]
            )
            commute_key = (
                row.revision.effective_at.date(),
                str(snapshot.get("origin_id", "")),
                str(snapshot.get("destination_id", "")),
            )
            is_new_commute = commute_key not in commute_days
            if tax_relevant and (
                derivation.classification != "commuting_allowance" or is_new_commute
            ):
                tax_amount = derivation.amount
                if derivation.classification == "commuting_allowance" and snapshot.get(
                    "transport_mode"
                ) not in {"private_car", "employer_car"}:
                    year = row.revision.effective_at.year
                    used = capped_commuting_totals.get(year, Decimal("0.00"))
                    tax_amount = min(tax_amount, max(Decimal("0.00"), Decimal("4500") - used))
                    capped_commuting_totals[year] = used + tax_amount
                if tax_amount:
                    tax_rows.append(
                        [
                            row.revision.effective_at.date().isoformat(),
                            str(row.event.pk),
                            derivation.classification,
                            str(derivation.distance_km),
                            derivation.rule_code,
                            tax_amount,
                        ]
                    )
                commute_days.add(commute_key)
            if employer_reimbursable and not snapshot.get("employer_paid"):
                personally_paid = Decimal(str(snapshot.get("personally_paid") or "0"))
                reimbursed = Decimal(str(snapshot.get("employer_reimbursement") or "0"))
                claim = max(Decimal("0.00"), personally_paid - reimbursed)
                if claim:
                    employer_rows.append(
                        [
                            row.revision.effective_at.date().isoformat(),
                            str(row.event.pk),
                            snapshot.get("transport_mode"),
                            "EUR",
                            claim,
                        ]
                    )
        elif row.event.event_type == "expense":
            amounts = expense_track_amounts(row.event, row.revision, as_of)
            expense_rows.append(
                [
                    str(row.event.pk),
                    row.revision.effective_at.isoformat(),
                    snapshot.get("category"),
                    snapshot.get("merchant"),
                    snapshot.get("amount"),
                    snapshot.get("currency"),
                    snapshot.get("employer_paid"),
                    tax_relevant,
                    employer_reimbursable,
                    snapshot.get("invoice_or_receipt_date"),
                    snapshot.get("payment_date"),
                    snapshot.get("payment_method"),
                    snapshot.get("vat_amount"),
                    snapshot.get("original_amount"),
                    snapshot.get("original_currency"),
                    snapshot.get("exchange_rate_to_eur"),
                    snapshot.get("reference"),
                    snapshot.get("supplier_address"),
                    snapshot.get("business_reason"),
                    snapshot.get("documentation_status"),
                    snapshot.get("description"),
                    snapshot.get("gross_amount_eur"),
                    snapshot.get("amount_personally_paid_eur"),
                    snapshot.get("employer_reimbursement_amount_eur"),
                    snapshot.get("professional_use_percentage"),
                    snapshot.get("justification"),
                ]
            )
            if amounts.tax_deduction:
                tax_rows.append(
                    [
                        row.revision.effective_at.date().isoformat(),
                        str(row.event.pk),
                        snapshot.get("category"),
                        snapshot.get("amount"),
                        "actual_expense",
                        amounts.tax_deduction,
                    ]
                )
            if amounts.employer_claim:
                employer_rows.append(
                    [
                        row.revision.effective_at.date().isoformat(),
                        str(row.event.pk),
                        snapshot.get("category"),
                        snapshot.get("currency"),
                        amounts.employer_claim,
                    ]
                )
        elif row.event.event_type == "external_activity":
            external_activity_rows.append(
                [
                    str(row.event.pk),
                    row.revision.effective_at.isoformat(),
                    snapshot.get("start_at"),
                    snapshot.get("end_at"),
                    snapshot.get("destination_name"),
                    snapshot.get("destination_locality"),
                    snapshot.get("departure_context"),
                    snapshot.get("return_context"),
                    snapshot.get("overnight"),
                    snapshot.get("client"),
                    snapshot.get("project"),
                    snapshot.get("purpose"),
                    _canonical(snapshot.get("journey_leg_ids", [])),
                    _canonical(snapshot.get("provided_meals", {})),
                    _canonical(snapshot.get("provided_meal_copayments", {})),
                    snapshot.get("employer_per_diem_reimbursement"),
                    snapshot.get("three_month_limit_applies"),
                    tax_relevant,
                    employer_reimbursable,
                ]
            )
            calculation = PerDiemCalculation.objects.filter(
                activity_event=row.event, input_revision=row.revision
            ).first()
            if calculation is not None and calculation.complete and calculation.total:
                if tax_relevant:
                    tax_rows.append(
                        [
                            row.revision.effective_at.date().isoformat(),
                            str(row.event.pk),
                            "meal_per_diem",
                            "derived per-diem",
                            ",".join(calculation.rule_codes),
                            calculation.total,
                        ]
                    )
                if employer_reimbursable:
                    employer_rows.append(
                        [
                            row.revision.effective_at.date().isoformat(),
                            str(row.event.pk),
                            "meal_per_diem",
                            "EUR",
                            calculation.total,
                        ]
                    )

    _write_sheet(
        workbook,
        "Journeys",
        [
            "event_id",
            "effective_at",
            "origin",
            "destination",
            "transport_mode",
            "actual_km",
            "tax_classification",
            "rule_code",
            "derived_amount",
            "complete",
            "covered_by_pass",
            "rail_pass_name",
            "total_fare",
            "personally_paid",
            "employer_reimbursement",
            "employer_paid",
            "payer_description",
        ],
        journey_rows,
    )
    _write_sheet(
        workbook,
        "ExternalActivities",
        [
            "event_id",
            "effective_at",
            "start_at",
            "end_at",
            "destination",
            "locality",
            "departed_from",
            "returned_to",
            "overnight",
            "client",
            "project",
            "purpose",
            "journey_leg_ids_json",
            "provided_meals_json",
            "meal_copayments_json",
            "employer_per_diem_reimbursement",
            "three_month_limit_applies",
            "tax_relevant",
            "employer_reimbursable",
        ],
        external_activity_rows,
    )
    _write_sheet(
        workbook,
        "Expenses",
        [
            "event_id",
            "effective_at",
            "category",
            "merchant",
            "amount",
            "currency",
            "employer_paid",
            "tax_relevant",
            "employer_reimbursable",
            "invoice_or_receipt_date",
            "payment_date",
            "payment_method",
            "vat_amount",
            "original_amount",
            "original_currency",
            "exchange_rate_to_eur",
            "reference",
            "supplier_address",
            "business_reason",
            "documentation_status",
            "description",
            "gross_amount_eur",
            "amount_personally_paid_eur",
            "employer_reimbursement_amount_eur",
            "professional_use_percentage",
            "justification",
        ],
        expense_rows,
    )

    selected_ids = [row.event.pk for row in selected]
    per_diem_rows = [
        [
            str(calculation.activity_event_id),
            str(calculation.input_revision_id),
            _canonical(calculation.rule_codes),
            _canonical(calculation.daily_amounts),
            calculation.total,
            calculation.complete,
            calculation.derivation_hash,
        ]
        for calculation in PerDiemCalculation.objects.filter(activity_event_id__in=selected_ids)
        .order_by("generated_at", "id")
    ]
    _write_sheet(
        workbook,
        "PerDiem",
        [
            "activity_event_id",
            "input_revision_id",
            "rule_codes",
            "daily_amounts",
            "total",
            "complete",
            "derivation_hash",
        ],
        per_diem_rows,
    )
    _write_sheet(
        workbook,
        "TaxOutput",
        ["date", "event_id", "category", "basis", "rule_code", "amount_eur"],
        tax_rows,
    )
    tax_sheet = workbook["TaxOutput"]
    tax_total_row = len(tax_rows) + 2
    tax_sheet.cell(tax_total_row, 5, "TOTAL")
    tax_sheet.cell(
        tax_total_row,
        6,
        f"=SUM(F2:F{len(tax_rows) + 1})" if tax_rows else Decimal("0.00"),
    )

    _write_sheet(
        workbook,
        "EmployerClaims",
        ["date", "event_id", "category", "currency", "claim_amount"],
        employer_rows,
    )
    employer_sheet = workbook["EmployerClaims"]
    employer_total_row = len(employer_rows) + 2
    employer_sheet.cell(employer_total_row, 4, "TOTAL")
    employer_sheet.cell(
        employer_total_row,
        5,
        f"=SUM(E2:E{len(employer_rows) + 1})" if employer_rows else Decimal("0.00"),
    )
    _write_sheet(
        workbook,
        "Reconciliation",
        ["metric", "value"],
        [
            ["selected_events", len(selected)],
            ["incomplete_events", sum(not row.revision.complete for row in selected)],
            [
                "tax_track_events",
                sum(
                    bool(row.revision.snapshot.get("tax_relevant", row.event.tax_relevant))
                    for row in selected
                ),
            ],
            [
                "employer_track_events",
                sum(
                    bool(
                        row.revision.snapshot.get(
                            "employer_reimbursable", row.event.employer_reimbursable
                        )
                    )
                    for row in selected
                ),
            ],
            [
                "unmatched_receipts",
                sum(
                    row.event.event_type == "receipt_only"
                    and row.revision.snapshot.get("reconciliation_status", "unmatched")
                    == "unmatched"
                    for row in selected
                ),
            ],
            ["tax_total_eur", f"=TaxOutput!F{tax_total_row}"],
            ["employer_claim_total_eur", f"=EmployerClaims!E{employer_total_row}"],
        ],
    )

    links = AttachmentLink.objects.filter(event_id__in=selected_ids).select_related("attachment")
    attachment_rows = [
        [
            str(link.attachment_id),
            str(link.event_id),
            link.link_type,
            link.attachment.original_filename,
            link.attachment.detected_format,
            link.attachment.size_bytes,
            link.attachment.sha256,
            link.attachment.relative_original_path,
        ]
        for link in links.order_by("event_id", "attachment_id", "link_type")
    ]
    _write_sheet(
        workbook,
        "Attachments",
        [
            "attachment_id",
            "event_id",
            "link_type",
            "original_filename",
            "format",
            "size_bytes",
            "sha256",
            "relative_path",
        ],
        attachment_rows,
    )

    revision_rows = [
        [
            str(revision.event_id),
            str(revision.pk),
            revision.revision_number,
            str(revision.parent_revision_id or ""),
            revision.effective_at.isoformat(),
            revision.recorded_at.isoformat(),
            revision.complete,
            revision.deleted,
            revision.comment,
            _canonical(revision.snapshot),
            revision.previous_audit_hash,
            revision.audit_hash,
        ]
        for revision in EventRevision.objects.filter(
            event_id__in=selected_ids, recorded_at__lte=as_of
        )
        .order_by("event_id", "revision_number")
    ]
    _write_sheet(
        workbook,
        "Revisions",
        [
            "event_id",
            "revision_id",
            "revision_number",
            "parent_revision_id",
            "effective_at",
            "recorded_at",
            "complete",
            "deleted",
            "comment",
            "snapshot_json",
            "previous_audit_hash",
            "audit_hash",
        ],
        revision_rows,
    )

    rule_rows = [
        [
            rule.code,
            rule.jurisdiction,
            rule.rule_type,
            rule.effective_from.isoformat(),
            rule.effective_to.isoformat() if rule.effective_to else "",
            _canonical(rule.values),
            rule.source_citation,
            rule.source_url,
        ]
        for rule in _tax_rule_rows(start, end)
    ]
    _write_sheet(
        workbook,
        "TaxRules",
        [
            "code",
            "jurisdiction",
            "rule_type",
            "effective_from",
            "effective_to",
            "values_json",
            "source_citation",
            "source_url",
        ],
        rule_rows,
    )
    _write_sheet(
        workbook,
        "README",
        ["key", "value"],
        [
            ["application", "WorkLedger"],
            ["range_start", start.isoformat()],
            ["range_end", end.isoformat()],
            ["as_of", as_of.isoformat()],
            ["event_count", len(selected)],
            ["notes", "User facts and derived tax outputs are separate; rule sources are listed."],
        ],
    )

    destination.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory() as temporary:
        raw_path = Path(temporary) / "raw.xlsx"
        workbook.save(raw_path)
        _normalized_zip(raw_path, destination)
    return destination


SQL_SCHEMA = """
PRAGMA foreign_keys=ON;
CREATE TABLE metadata (key TEXT PRIMARY KEY, value TEXT NOT NULL);
CREATE TABLE events (id TEXT PRIMARY KEY, event_type TEXT NOT NULL, effective_at TEXT NOT NULL,
 recorded_at TEXT NOT NULL, revision_number INTEGER NOT NULL, complete INTEGER NOT NULL,
 tax_relevant INTEGER NOT NULL, employer_reimbursable INTEGER NOT NULL, snapshot_json TEXT NOT NULL,
 audit_hash TEXT NOT NULL);
CREATE TABLE revisions (id TEXT PRIMARY KEY, event_id TEXT NOT NULL,
 revision_number INTEGER NOT NULL, parent_revision_id TEXT, effective_at TEXT NOT NULL,
 recorded_at TEXT NOT NULL, complete INTEGER NOT NULL,
 deleted INTEGER NOT NULL, comment TEXT NOT NULL, snapshot_json TEXT NOT NULL,
 previous_audit_hash TEXT NOT NULL, audit_hash TEXT NOT NULL);
CREATE TABLE attachments (id TEXT PRIMARY KEY, original_filename TEXT NOT NULL,
 detected_format TEXT NOT NULL,
 size_bytes INTEGER NOT NULL, sha256 TEXT NOT NULL, relative_original_path TEXT NOT NULL,
 relative_preview_path TEXT NOT NULL, relative_thumbnail_path TEXT NOT NULL);
CREATE TABLE attachment_links (attachment_id TEXT NOT NULL, event_id TEXT NOT NULL,
 link_type TEXT NOT NULL,
 PRIMARY KEY (attachment_id,event_id,link_type));
CREATE TABLE tax_rules (code TEXT PRIMARY KEY, jurisdiction TEXT NOT NULL, rule_type TEXT NOT NULL,
 effective_from TEXT NOT NULL, effective_to TEXT, values_json TEXT NOT NULL,
 source_citation TEXT NOT NULL, source_url TEXT NOT NULL);
"""


def build_range_sqlite(
    start: date,
    end: date,
    as_of: datetime,
    destination: Path,
    events: list[ExportEvent] | None = None,
) -> Path:
    selected = events if events is not None else _events_as_of(start, end, as_of)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.unlink(missing_ok=True)
    selected_ids = [row.event.pk for row in selected]
    with sqlite3.connect(destination) as connection:
        connection.executescript(SQL_SCHEMA)
        connection.executemany(
            "INSERT INTO metadata VALUES (?,?)",
            [
                ("application", "WorkLedger"),
                ("range_start", start.isoformat()),
                ("range_end", end.isoformat()),
                ("as_of", as_of.isoformat()),
            ],
        )
        connection.executemany(
            "INSERT INTO events VALUES (?,?,?,?,?,?,?,?,?,?)",
            [
                (
                    str(row.event.pk),
                    row.event.event_type,
                    row.revision.effective_at.isoformat(),
                    row.revision.recorded_at.isoformat(),
                    row.revision.revision_number,
                    row.revision.complete,
                    bool(
                        row.revision.snapshot.get("tax_relevant", row.event.tax_relevant)
                    ),
                    bool(
                        row.revision.snapshot.get(
                            "employer_reimbursable", row.event.employer_reimbursable
                        )
                    ),
                    _canonical(row.revision.snapshot),
                    row.revision.audit_hash,
                )
                for row in selected
            ],
        )
        revisions = EventRevision.objects.filter(
            event_id__in=selected_ids, recorded_at__lte=as_of
        ).order_by("event_id", "revision_number")
        connection.executemany(
            "INSERT INTO revisions VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            [
                (
                    str(revision.pk),
                    str(revision.event_id),
                    revision.revision_number,
                    str(revision.parent_revision_id or ""),
                    revision.effective_at.isoformat(),
                    revision.recorded_at.isoformat(),
                    revision.complete,
                    revision.deleted,
                    revision.comment,
                    _canonical(revision.snapshot),
                    revision.previous_audit_hash,
                    revision.audit_hash,
                )
                for revision in revisions
            ],
        )
        links = list(
            AttachmentLink.objects.filter(event_id__in=selected_ids)
            .select_related("attachment")
            .order_by("attachment_id", "event_id", "link_type")
        )
        attachments = {str(link.attachment_id): link.attachment for link in links}
        connection.executemany(
            "INSERT INTO attachments VALUES (?,?,?,?,?,?,?,?)",
            [
                (
                    attachment_id,
                    attachment.original_filename,
                    attachment.detected_format,
                    attachment.size_bytes,
                    attachment.sha256,
                    attachment.relative_original_path,
                    attachment.relative_preview_path,
                    attachment.relative_thumbnail_path,
                )
                for attachment_id, attachment in sorted(attachments.items())
            ],
        )
        connection.executemany(
            "INSERT INTO attachment_links VALUES (?,?,?)",
            [(str(link.attachment_id), str(link.event_id), link.link_type) for link in links],
        )
        connection.executemany(
            "INSERT INTO tax_rules VALUES (?,?,?,?,?,?,?,?)",
            [
                (
                    rule.code,
                    rule.jurisdiction,
                    rule.rule_type,
                    rule.effective_from.isoformat(),
                    rule.effective_to.isoformat() if rule.effective_to else None,
                    _canonical(rule.values),
                    rule.source_citation,
                    rule.source_url,
                )
                for rule in _tax_rule_rows(start, end)
            ],
        )
        connection.commit()
        connection.execute("VACUUM")
    return destination


def _write_deterministic_file(archive: zipfile.ZipFile, name: str, content: bytes) -> None:
    info = zipfile.ZipInfo(name, FIXED_ZIP_TIME)
    info.compress_type = zipfile.ZIP_DEFLATED
    info.create_system = 3
    info.external_attr = 0o100644 << 16
    archive.writestr(info, content, compress_type=zipfile.ZIP_DEFLATED)


def build_range_csv(
    start: date,
    end: date,
    as_of: datetime,
    destination: Path,
    selected: list[ExportEvent] | None = None,
) -> Path:
    rows = selected if selected is not None else _events_as_of(start, end, as_of)
    output = io.StringIO(newline="")
    writer = csv.writer(output, quoting=csv.QUOTE_ALL, lineterminator="\n")
    writer.writerow(
        [
            "event_id",
            "event_type",
            "effective_at",
            "recorded_at",
            "complete",
            "tax_relevant",
            "employer_reimbursable",
            "snapshot_json",
            "audit_hash",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                str(row.event.pk),
                row.event.event_type,
                row.revision.effective_at.isoformat(),
                row.revision.recorded_at.isoformat(),
                row.revision.complete,
                bool(row.revision.snapshot.get("tax_relevant", row.event.tax_relevant)),
                bool(
                    row.revision.snapshot.get(
                        "employer_reimbursable", row.event.employer_reimbursable
                    )
                ),
                _canonical(row.revision.snapshot),
                row.revision.audit_hash,
            ]
        )
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(output.getvalue(), encoding="utf-8", newline="")
    return destination


def build_range_json(
    start: date,
    end: date,
    as_of: datetime,
    destination: Path,
    selected: list[ExportEvent] | None = None,
) -> Path:
    rows = selected if selected is not None else _events_as_of(start, end, as_of)
    payload = {
        "application": "WorkLedger",
        "range_start": start.isoformat(),
        "range_end": end.isoformat(),
        "as_of": as_of.isoformat(),
        "events": [
            {
                "event_id": str(row.event.pk),
                "event_type": row.event.event_type,
                "revision_id": str(row.revision.pk),
                "effective_at": row.revision.effective_at.isoformat(),
                "recorded_at": row.revision.recorded_at.isoformat(),
                "complete": row.revision.complete,
                "deleted": row.revision.deleted,
                "snapshot": row.revision.snapshot,
                "audit_hash": row.revision.audit_hash,
            }
            for row in rows
        ],
    }
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
        encoding="utf-8",
    )
    return destination


def build_full_zip(
    start: date,
    end: date,
    as_of: datetime,
    destination: Path,
    selected: list[ExportEvent] | None = None,
) -> Path:
    selected = selected if selected is not None else _events_as_of(start, end, as_of)
    selected_ids = [row.event.pk for row in selected]
    links = list(
        AttachmentLink.objects.filter(event_id__in=selected_ids)
        .select_related("attachment")
        .order_by("attachment_id", "event_id")
    )
    with tempfile.TemporaryDirectory() as temporary:
        work = Path(temporary)
        xlsx = build_range_xlsx(start, end, as_of, work / "workledger.xlsx", selected)
        sqlite = build_range_sqlite(start, end, as_of, work / "workledger.sqlite3", selected)
        json_export = build_range_json(start, end, as_of, work / "data.json", selected)
        files: dict[str, bytes] = {
            "workledger.xlsx": xlsx.read_bytes(),
            "workledger.sqlite3": sqlite.read_bytes(),
            "data.json": json_export.read_bytes(),
        }
        workbook = load_workbook(xlsx, data_only=False, read_only=True)
        csv_sheets = {
            "events.csv": "Events",
            "journeys.csv": "Journeys",
            "external-activities.csv": "ExternalActivities",
            "derived-days.csv": "PerDiem",
            "expenses.csv": "Expenses",
            "reimbursements.csv": "EmployerClaims",
            "attachments.csv": "Attachments",
            "revisions.csv": "Revisions",
        }
        dictionary_lines = ["# WorkLedger data dictionary", ""]
        for filename, sheet_name in csv_sheets.items():
            buffer = io.StringIO(newline="")
            writer = csv.writer(buffer, quoting=csv.QUOTE_ALL, lineterminator="\n")
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            for row in rows:
                writer.writerow(["" if value is None else value for value in row])
            files[filename] = buffer.getvalue().encode()
            headers = [str(value) for value in rows[0]] if rows else []
            dictionary_lines.extend(
                [f"## {filename}", "", ", ".join(f"`{header}`" for header in headers), ""]
            )
        files["readme.md"] = (
            "# WorkLedger complete export\n\n"
            f"Range: {start.isoformat()} through {end.isoformat()}\n\n"
            "Files are deterministic snapshots as of " + as_of.isoformat() + ".\n"
        ).encode()
        files["data-dictionary.md"] = "\n".join(dictionary_lines).encode()
        unresolved = [row for row in selected if not row.revision.complete]
        files["unresolved-items.md"] = (
            "# Unresolved items\n\n"
            + "".join(
                f"- `{row.event.pk}` {row.event.event_type} at "
                f"{row.revision.effective_at.isoformat()}\n"
                for row in unresolved
            )
            + ("- None\n" if not unresolved else "")
        ).encode()
        seen: set[str] = set()
        for link in links:
            attachment: Attachment = link.attachment
            if str(attachment.pk) in seen:
                continue
            seen.add(str(attachment.pk))
            original_name = f"attachments/originals/{attachment.pk}_{attachment.original_filename}"
            files[original_name] = attachment.original_path.read_bytes()
            if attachment.relative_preview_path and attachment.preview_path.exists():
                preview_name = f"attachments/previews/{attachment.pk}.jpg"
                files[preview_name] = attachment.preview_path.read_bytes()
            if attachment.relative_thumbnail_path and attachment.thumbnail_path.exists():
                thumbnail_name = f"attachments/thumbnails/{attachment.pk}.jpg"
                files[thumbnail_name] = attachment.thumbnail_path.read_bytes()
        manifest = {
            "application": "WorkLedger",
            "range_start": start.isoformat(),
            "range_end": end.isoformat(),
            "as_of": as_of.isoformat(),
            "files": {
                name: {"sha256": hashlib.sha256(content).hexdigest(), "size": len(content)}
                for name, content in sorted(files.items())
            },
        }
        files["manifest.json"] = json.dumps(
            manifest, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        ).encode()
        destination.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(
            destination, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as archive:
            for name, content in sorted(files.items()):
                _write_deterministic_file(archive, name, content)
    return destination


def generate_export(
    *, start: date, end: date, kind: str, as_of: datetime
) -> ExportArtifact:
    from .models import ExportArtifact

    export_id = uuid.uuid4()
    suffixes: dict[str, str] = {
        ExportArtifact.Kind.XLSX.value: ".xlsx",
        ExportArtifact.Kind.CSV.value: ".csv",
        ExportArtifact.Kind.JSON.value: ".json",
        ExportArtifact.Kind.SQLITE.value: ".sqlite3",
        ExportArtifact.Kind.FULL_ZIP.value: ".zip",
    }
    if kind not in suffixes:
        raise ValueError(f"Unsupported export kind: {kind}")
    relative_path = f"exports/{start}_{end}_{export_id}{suffixes[kind]}"
    destination = Path(settings.DATA_DIR) / relative_path
    builders: dict[str, Callable[..., Path]] = {
        ExportArtifact.Kind.XLSX.value: build_range_xlsx,
        ExportArtifact.Kind.CSV.value: build_range_csv,
        ExportArtifact.Kind.JSON.value: build_range_json,
        ExportArtifact.Kind.SQLITE.value: build_range_sqlite,
        ExportArtifact.Kind.FULL_ZIP.value: build_full_zip,
    }
    builders[kind](start, end, as_of, destination)
    content = destination.read_bytes()
    try:
        return ExportArtifact.objects.create(
            id=export_id,
            kind=kind,
            range_start=start,
            range_end=end,
            as_of=as_of,
            relative_path=relative_path,
            sha256=hashlib.sha256(content).hexdigest(),
            size_bytes=len(content),
        )
    except Exception:
        destination.unlink(missing_ok=True)
        raise
