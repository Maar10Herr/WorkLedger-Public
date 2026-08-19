from __future__ import annotations

import hashlib
import json
import tempfile
import zipfile
from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from uuid import UUID

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook

from apps.evidence.models import AttachmentLink
from apps.expenses.services import expense_track_amounts
from apps.ledger.models import Event
from apps.ledger.services import create_event
from apps.taxes.models import PerDiemCalculation

from .models import EmployerPackage, PackageEvent, PackageStatusChange
from .services import (
    _normalized_zip,
    _write_deterministic_file,
    _write_sheet,
)


@transaction.atomic
def create_package(
    *,
    name: str,
    period_start: date,
    period_end: date,
    event_ids: Iterable[object],
) -> EmployerPackage:
    requested_ids = list(dict.fromkeys(event_ids))
    try:
        normalized_ids = [UUID(str(event_id)) for event_id in requested_ids]
    except (TypeError, ValueError, AttributeError) as exc:
        raise ValidationError("One or more selected events are invalid") from exc
    if not normalized_ids:
        raise ValidationError("Select at least one reimbursable event")
    events = list(
        Event.objects.filter(
            pk__in=normalized_ids,
            employer_reimbursable=True,
            current_revision__effective_at__date__range=(period_start, period_end),
        )
        .select_related("current_revision")
        .order_by("id")
    )
    eligible_events = [
        event
        for event in events
        if event.current_revision is not None
        and not event.current_revision.deleted
        and event.current_revision.complete
    ]
    if len(eligible_events) != len(normalized_ids):
        raise ValidationError(
            "Every selected event must exist, be complete, reimbursable, and inside the period"
        )
    package = EmployerPackage.objects.create(
        name=name, period_start=period_start, period_end=period_end
    )
    claim_total = Decimal("0.00")
    for event in eligible_events:
        revision = event.current_revision
        assert revision is not None
        amount = Decimal("0.00")
        if event.event_type == "expense":
            amount = expense_track_amounts(event, revision).employer_claim
        elif event.event_type == "journey":
            personally_paid = Decimal(str(revision.snapshot.get("personally_paid") or "0"))
            reimbursed = Decimal(
                str(revision.snapshot.get("employer_reimbursement") or "0")
            )
            amount = max(Decimal("0.00"), personally_paid - reimbursed)
        elif event.event_type == "external_activity":
            calculation = PerDiemCalculation.objects.filter(
                activity_event=event, input_revision=revision, complete=True
            ).first()
            amount = calculation.total if calculation is not None else Decimal("0.00")
        else:
            amount = Decimal(str(revision.snapshot.get("employer_claim_amount", "0")))
        PackageEvent.objects.create(
            package=package,
            event=event,
            included_revision=revision,
            claimed_amount=amount,
        )
        claim_total += amount
    package.claim_amount = claim_total.quantize(Decimal("0.01"))
    package.save(update_fields=["claim_amount"])
    return package


VALID_TRANSITIONS: dict[str, set[str]] = {
    EmployerPackage.Status.DRAFT.value: {EmployerPackage.Status.SUBMITTED.value},
    EmployerPackage.Status.SUBMITTED.value: {
        EmployerPackage.Status.REIMBURSED.value,
        EmployerPackage.Status.REJECTED.value,
    },
    EmployerPackage.Status.REJECTED.value: {EmployerPackage.Status.SUBMITTED.value},
    EmployerPackage.Status.REIMBURSED.value: set(),
}


def _build_claim_workbook(
    package: EmployerPackage, memberships: list[PackageEvent], destination: Path
) -> Path:
    workbook = Workbook()
    active = workbook.active
    assert active is not None
    workbook.remove(active)
    workbook.properties.creator = "WorkLedger"
    workbook.properties.lastModifiedBy = "WorkLedger"
    workbook.properties.created = datetime(1980, 1, 1)
    workbook.properties.modified = datetime(1980, 1, 1)
    rows: list[list[object]] = []
    for membership in memberships:
        revision = membership.included_revision
        snapshot = revision.snapshot
        rows.append(
            [
                revision.effective_at.date().isoformat(),
                str(membership.event_id),
                membership.event.get_event_type_display(),
                snapshot.get("category", ""),
                snapshot.get("merchant", ""),
                snapshot.get("purpose", snapshot.get("note", "")),
                membership.claimed_amount,
                package.currency,
            ]
        )
    _write_sheet(
        workbook,
        "Claim",
        [
            "date",
            "event_id",
            "type",
            "category",
            "merchant",
            "purpose",
            "amount",
            "currency",
        ],
        rows,
    )
    claim_sheet = workbook["Claim"]
    total_row = len(rows) + 2
    claim_sheet.cell(total_row, 6, "TOTAL")
    claim_sheet.cell(total_row, 7, f"=SUM(G2:G{len(rows) + 1})")
    links = (
        AttachmentLink.objects.filter(
            event_id__in=[membership.event_id for membership in memberships]
        )
        .select_related("attachment")
        .order_by("event_id", "attachment_id")
    )
    _write_sheet(
        workbook,
        "Evidence",
        ["event_id", "filename", "sha256", "size_bytes"],
        [
            [
                str(link.event_id),
                link.attachment.original_filename,
                link.attachment.sha256,
                link.attachment.size_bytes,
            ]
            for link in links
        ],
    )
    _write_sheet(
        workbook,
        "README",
        ["field", "value"],
        [
            ["package", package.name],
            ["period", f"{package.period_start} to {package.period_end}"],
            ["status", package.status],
            [
                "scope",
                "Employer reimbursement only; tax-deduction output is intentionally excluded.",
            ],
        ],
    )
    temporary = destination.with_suffix(".raw.xlsx")
    workbook.save(temporary)
    _normalized_zip(temporary, destination)
    temporary.unlink(missing_ok=True)
    return destination


@transaction.atomic
def update_package_status(
    package: EmployerPackage, new_status: str, *, note: str = ""
) -> PackageStatusChange:
    locked = EmployerPackage.objects.select_for_update().get(pk=package.pk)
    if new_status not in VALID_TRANSITIONS[locked.status]:
        raise ValueError(f"Invalid package transition: {locked.status} → {new_status}")
    changed_at = timezone.now()
    status_event = create_event(
        event_type="reimbursement_update",
        effective_at=changed_at,
        snapshot={
            "package_id": str(locked.pk),
            "package_name": locked.name,
            "from_status": locked.status,
            "to_status": new_status,
            "claim_amount": str(locked.claim_amount),
            "currency": locked.currency,
            "note": note,
        },
        complete=True,
        employer_reimbursable=True,
        comment="Employer package status update",
    )
    change = PackageStatusChange.objects.create(
        package=locked,
        status_event=status_event,
        from_status=locked.status,
        to_status=new_status,
        note=note,
    )
    locked.status = new_status
    if new_status == EmployerPackage.Status.SUBMITTED:
        locked.submitted_at = changed_at
    if new_status in {EmployerPackage.Status.REIMBURSED, EmployerPackage.Status.REJECTED}:
        locked.resolved_at = changed_at
    locked.save(update_fields=["status", "submitted_at", "resolved_at"])
    if new_status in {
        EmployerPackage.Status.SUBMITTED,
        EmployerPackage.Status.REIMBURSED,
        EmployerPackage.Status.REJECTED,
    }:
        from apps.expenses.models import Expense
        from apps.expenses.services import update_reimbursement_status

        expense_status_by_package_status: dict[str, str] = {
            EmployerPackage.Status.SUBMITTED.value: Expense.ReimbursementStatus.SUBMITTED.value,
            EmployerPackage.Status.REIMBURSED.value: Expense.ReimbursementStatus.REIMBURSED.value,
            EmployerPackage.Status.REJECTED.value: Expense.ReimbursementStatus.REJECTED.value,
        }
        expense_status = expense_status_by_package_status[new_status]
        for expense in Expense.objects.filter(event__package_memberships__package=locked):
            if (
                new_status == EmployerPackage.Status.SUBMITTED
                and expense.reimbursement_status == Expense.ReimbursementStatus.REJECTED
            ):
                update_reimbursement_status(
                    expense,
                    Expense.ReimbursementStatus.READY.value,
                    note=f"Employer package {locked.pk} reopened",
                )
            update_reimbursement_status(
                expense, expense_status, note=f"Employer package {locked.pk}"
            )
    if new_status == EmployerPackage.Status.SUBMITTED:
        generate_package_zip(locked, as_of=changed_at)
    package.status = locked.status
    package.submitted_at = locked.submitted_at
    package.resolved_at = locked.resolved_at
    package.relative_package_path = locked.relative_package_path
    package.package_sha256 = locked.package_sha256
    return change


def generate_package_zip(package: EmployerPackage, *, as_of: datetime | None = None) -> Path:
    if (
        package.status != EmployerPackage.Status.DRAFT
        and package.relative_package_path
        and package.package_path.exists()
    ):
        return package.package_path
    frozen_at = as_of or timezone.now()
    memberships = list(
        package.package_events.select_related("event", "included_revision").order_by("event_id")
    )
    event_ids = {membership.event_id for membership in memberships}
    relative_path = f"employer-packages/{package.pk}_{package.status}.zip"
    destination = Path(settings.DATA_DIR) / relative_path
    destination.parent.mkdir(parents=True, exist_ok=True)
    links = list(
        AttachmentLink.objects.filter(event_id__in=event_ids)
        .select_related("attachment")
        .order_by("attachment_id", "event_id")
    )
    with tempfile.TemporaryDirectory() as temporary:
        claim_path = _build_claim_workbook(
            package, memberships, Path(temporary) / "claim.xlsx"
        )
        files: dict[str, bytes] = {"claim.xlsx": claim_path.read_bytes()}
        seen: set[object] = set()
        for link in links:
            attachment = link.attachment
            if attachment.pk in seen:
                continue
            seen.add(attachment.pk)
            files[
                f"receipts/{attachment.pk}_{attachment.original_filename}"
            ] = attachment.original_path.read_bytes()
        manifest = {
            "application": "WorkLedger",
            "package_id": str(package.pk),
            "name": package.name,
            "status": package.status,
            "period_start": package.period_start.isoformat(),
            "period_end": package.period_end.isoformat(),
            "claim_amount": str(package.claim_amount),
            "currency": package.currency,
            "as_of": frozen_at.isoformat(),
            "event_ids": sorted(str(event_id) for event_id in event_ids),
            "files": {
                name: {"sha256": hashlib.sha256(content).hexdigest(), "size": len(content)}
                for name, content in sorted(files.items())
            },
        }
        files["manifest.json"] = json.dumps(
            manifest, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        ).encode()
        with zipfile.ZipFile(
            destination, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as archive:
            for name, content in sorted(files.items()):
                _write_deterministic_file(archive, name, content)
    package.relative_package_path = relative_path
    package.package_sha256 = hashlib.sha256(destination.read_bytes()).hexdigest()
    package.save(update_fields=["relative_package_path", "package_sha256"])
    return destination
